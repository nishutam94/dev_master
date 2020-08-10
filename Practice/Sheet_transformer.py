import openpyxl as xl; 
print("hello")
filename1 ="/root/nishant/snap/MKL_un2.xlsx"
filename="/root/nishant/snap/MKL_un.xlsx"
wb1 = xl.load_workbook(filename) 
print(wb1)
ws1 = wb1.worksheets[0] 
wb = xl.Workbook()
sheet_obj = wb.active

def filldata(word,j):
	for k in range(len(word)):
		obj = sheet_obj.cell(row = j,column = k+1)
		obj.value = word[k]
	print("row : ", word, "Added")
		#print(j,k+1,word[k],"added")  

list_of_inter=[1,2,4,8,10,12,14,16,18,20,22,24]
mr = ws1.max_row 
mc = ws1.max_column 
print(mr)

row_count=0
for inter in list_of_inter:
	print(inter)
	row_count=row_count+1
	word = [] 
	print("***********************for inter : ",inter," : ********************")
	for i in range (1, mr + 1):
		row_inter=(ws1.cell(row = i, column=5).value)
		inter=int(inter)
		print(inter,row_inter)
		if (row_inter==inter):
			word.append(ws1.cell(row = i, column=9).value)
	print(word)			
	filldata(word,row_count)

#ws2.cell(row = i, column = j).value = c.value 
wb.save(str(filename1))
