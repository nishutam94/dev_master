from __future__ import print_function
from argparse import ArgumentParser
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import os
import re
import math
import json
wb = openpyxl.Workbook()
sheet_obj = wb.active
colour = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')

def reguler_expression(regx,i):
	with open('config.json', 'r') as f:
		config_dict = json.load(f)
		print(config_dict)
		conf = list(config_dict)
		print(config_dict["wnd"][i][regx])
	return config_dict["wnd"][i][regx]
	#reg={ "throughput":r"Images per second: ([0-9.]*)", "accuracy":r"Accuracy: ([0-9.]*)%"}
	#reg={ "throughput": r"Throughput is \(records/sec\)  :  ([0-9.]*)", "accuracy": r"Classification accuracy \(%\)  :  ([0-9.]*)", "latency":r"'latency"}
	#return reg.get(regx)

def validate(match,regx,test):
	if(not match):
		print("No Match for",regx)
		match = " "
	return match[0]	

def filldata(word,j):
    	for k in range(len(word)):
        	obj = sheet_obj.cell(row = j,column = k+1)
        	obj.value = word[k]
    	print("row : ", word, "Added")
        #print(j,k+1,word[k],"added")

def tamplate():
	#header=["Framework","Branch","Topology","Model","test","bs","batchsize","i","itration","type","plateform","Throughput","Latency","Accuracy","Machine"]
	header=["Machine","Framework","Branch","Topology","Model","test","batchsize","itration","type","plateform","Throughput","Latency","Accuracy"]
	filldata(header,1)

def recursive(path,file,j):
    print("log name: ",file)
    words = file.split('.')
    word = []
    word = words[0].split('_')    
    with open (path + file, "r+") as myfile:
        data = myfile.read()
        #print(data)
        tests =["throughput","accuracy","latency"]
        #tests =["throughput"]
        testvalue=[]
        for test in tests:
            i=0
            regx = reguler_expression(test,i)
            match = re.findall(regx,data)
            print(match)
            word.append(validate(match,regx,test))
            i=i+1
    print(j)        
    filldata(word,j)

def main():
	parser = ArgumentParser()
	parser.add_argument("-m", "--machine", help="Specify Machine Name",default="SDP5")
	parser.add_argument("-p", "--path", help="Specify Path of logs",default="/root/nishant/logs/")
	parser.add_argument("-o", "--output", help="Specify name of summary",default="summary.xlsx")
	args = parser.parse_args()
	machine = args.machine
	path = args.path
	output=args.output
	tamplate()
	filelist = os.listdir(path)
	j=2
	for file in filelist:
		j=j+1
		word = recursive(path,file,j)
	wb.save(output)		

if __name__ == "__main__":
	main()